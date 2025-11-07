# start.py
from openpyxl import Workbook
import os

FILE = "snack.xlsx"

name = input("이름 입력: ")

if not os.path.exists(FILE):
    wb = Workbook()
else:
    from openpyxl import load_workbook
    wb = load_workbook(FILE)

if name not in wb.sheetnames:
    ws = wb.create_sheet(name)
    ws["A1"], ws["B1"] = "Snack", "Count"
    ws["A2"], ws["B2"] = "초기자금", 5

wb.save(FILE)
print("✅ 저장 완료! 이제 main.py 실행.")